"""
Excel 导出工具
"""
import io
from datetime import datetime
from typing import Optional, List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """Excel 导出工具类"""
    
    # 默认样式
    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    
    BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    @classmethod
    def export_dataframe(
        cls,
        df: pd.DataFrame,
        sheet_name: str = "Sheet1",
        title: Optional[str] = None
    ) -> bytes:
        """
        将 DataFrame 导出为 Excel 字节流
        
        Args:
            df: 要导出的 DataFrame
            sheet_name: 工作表名称
            title: 可选的标题行
            
        Returns:
            bytes: Excel 文件字节流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        start_row = 1
        
        # 添加标题
        if title:
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            start_row = 3
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = cls.BORDER
                
                # 表头样式
                if r_idx == start_row:
                    cell.font = cls.HEADER_FONT
                    cell.fill = cls.HEADER_FILL
                    cell.alignment = cls.HEADER_ALIGNMENT
                else:
                    # 数据行样式
                    if isinstance(value, (int, float)):
                        cell.alignment = cls.NUMBER_ALIGNMENT
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = cls.CELL_ALIGNMENT
        
        # 自动调整列宽
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(df.columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    # 跳过合并单元格
                    if hasattr(cell, 'value') and cell.value:
                        cell_length = len(str(cell.value))
                        # 中文字符按 2 倍宽度计算
                        chinese_count = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                        cell_length += chinese_count
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @classmethod
    def export_multiple_sheets(
        cls,
        sheets: Dict[str, pd.DataFrame],
        title: Optional[str] = None
    ) -> bytes:
        """
        导出多个工作表到一个 Excel 文件
        
        Args:
            sheets: {工作表名称: DataFrame} 字典
            title: 可选的每个工作表的标题
            
        Returns:
            bytes: Excel 文件字节流
        """
        wb = Workbook()
        
        # 删除默认的工作表
        default_sheet = wb.active
        
        for idx, (sheet_name, df) in enumerate(sheets.items()):
            if idx == 0:
                ws = default_sheet
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            start_row = 1
            
            # 添加标题
            if title:
                ws.cell(row=1, column=1, value=f"{title} - {sheet_name}")
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                if len(df.columns) > 0:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                start_row = 3
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = cls.BORDER
                    
                    if r_idx == start_row:
                        cell.font = cls.HEADER_FONT
                        cell.fill = cls.HEADER_FILL
                        cell.alignment = cls.HEADER_ALIGNMENT
                    else:
                        if isinstance(value, (int, float)):
                            cell.alignment = cls.NUMBER_ALIGNMENT
                            if isinstance(value, float):
                                cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = cls.CELL_ALIGNMENT
            
            # 自动调整列宽
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(start_row, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            chinese_count = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                            cell_length += chinese_count
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def get_filename(prefix: str, suffix: str = "xlsx") -> str:
        """
        生成带时间戳的文件名
        
        Args:
            prefix: 文件名前缀
            suffix: 文件扩展名
            
        Returns:
            str: 文件名
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{suffix}"
